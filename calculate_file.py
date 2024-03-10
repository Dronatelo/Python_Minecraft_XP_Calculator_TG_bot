import openpyxl
import json
from json.decoder import JSONDecodeError
from settings.file_settings import data_items_path, now_file_name_path, excel_files_path

def get_file_now_name():
    with open(now_file_name_path) as file_nfnp:
        first_line = file_nfnp.readline().strip()
        return first_line

def del_line():
    with open(now_file_name_path) as file_nfnp:
        first_line = file_nfnp.readline()
    with open(now_file_name_path, 'w') as file_nfnp:
        file_nfnp.writelines(first_line[100:])

def add_line(name_file):
    with open(now_file_name_path, 'w') as file_nfnp:
        file_nfnp.write(name_file)

def xlms_path_now():
    last_name = get_file_now_name()
    return f'{excel_files_path}{last_name}'

def collect_items_data(path_json):
    
    def get_data_excel():
        data_items_dict = {}
        bb = []
        for row in range(2, sheet.max_row+1):
            data_items_dict = {
                "id_bot":sheet[row][0].value,
                "name_item":sheet[row][1].value,
                "craft_xp":sheet[row][2].value,
                "smit_xp":sheet[row][3].value,
                "magic_xp":sheet[row][4].value,
                "create_xp":sheet[row][5].value
            }

            bb.append(data_items_dict)

            with open(path_json,'w',encoding="utf-8") as item_json:
                json.dump(bb,item_json,indent=2,ensure_ascii=False)
        print("easy_data")
         
    file_data = openpyxl.open(xlms_path_now(),read_only=False)
    sheet = file_data.active
    
    with open(path_json, 'r',encoding="utf-8") as f:
        try:
            un_data = json.load(f)
            if un_data == "[]":
                get_data_excel()
                
        except JSONDecodeError:
            with open(path_json, 'w') as fw:
                fw.write(json.dumps([]))
    get_data_excel()
            
#id_item, amount - количество, path - путь до файла
def calculate_data(path_json,id_item,amount):
    collect_items_data(data_items_path)
    with open(path_json, 'r',encoding="utf-8") as uncoding:
        un_data = json.load(uncoding)

    for i in un_data:
        if i['id_bot']==id_item:
            return (i['craft_xp']*amount,i['smit_xp']*amount,i["magic_xp"]*amount,i["create_xp"]*amount)

def output_data(path_json):
    collect_items_data(data_items_path)
    with open(path_json, 'r',encoding="utf-8") as uncoding:
        un_data = json.load(uncoding)

    output =''
    
    for i in un_data:
        
        output+= "ID Bot: "f'{i["id_bot"]}'+"\n"+"Name: "f'{i["name_item"]}'+"\n"
        output+="\n"
        if i["craft_xp"] == 0:
            pass
        else:
            output+="Craft XP: "f'{i["craft_xp"]}'+"\n"
        if i["smit_xp"]==0:
            pass
        else:
            output+="Smit XP: "f'{i["smit_xp"]}'+"\n"
        if i["magic_xp"]==0:
            pass
        else:
            output+="Magic XP: "f'{i["magic_xp"]}'+"\n"
        if i["create_xp"]==0:
            pass
        else:
            output+="Create XP: "f'{i["create_xp"]}'+"\n"
            
        output+="-------------------------"+"\n"
        # output+=("ID Bot: "+i['id_bot']+'\n'+
        # 'Name: '+i["name_item"]+"\n"+
        # "Craft XP: "f'{i["craft_xp"]}'+"\n"+
        # "Smit XP: "f'{i["smit_xp"]}'+"\n"+
        # "Magic XP: "f'{i["magic_xp"]}'+"\n"+
        # "Create XP: "f'{i["create_xp"]}'+"\n"+
        # "------------------------"+"\n")
        
    return output

def add_data(id_bot, name_item, craft_xp,smit_xp):
    fi_data = openpyxl.open(xlms_path_now(),read_only=False)
    sheets = fi_data.active

    n = sheets.max_row+1
    sheets[n][0].value = id_bot
    sheets[n][1].value = name_item
    sheets[n][2].value = craft_xp
    sheets[n][3].value = smit_xp

    fi_data.save(xlms_path_now())
    fi_data.close()

def find_data_id(id):
    fi_data = openpyxl.open(xlms_path_now(),read_only=False)
    sheets = fi_data.active
    for i in range(2,sheets.max_row+1):
        if sheets[i][0].value == id:
            return i

def find_data_name(id):
    fi_data = openpyxl.open(xlms_path_now(),read_only=False)
    sheets = fi_data.active
    for i in range(2,sheets.max_row+1):
        if sheets[i][0].value == id:
            return sheets[i][1].value

def find_data_id_bot(id):
    fi_data = openpyxl.open(xlms_path_now(),read_only=False)
    sheets = fi_data.active
    for i in range(2,sheets.max_row+1):
        if sheets[i][0].value == id:
            return sheets[i][0].value

def edit_data(n,id_bot, name_item, craft_xp,smit_xp):
    fi_data = openpyxl.open(xlms_path_now(),read_only=False)
    sheets = fi_data.active

    sheets[n][0].value = id_bot
    sheets[n][1].value = name_item
    sheets[n][2].value = craft_xp
    sheets[n][3].value = smit_xp

    fi_data.save(xlms_path_now())
    fi_data.close()

def delete_data(n):
    fixe_data = openpyxl.open(xlms_path_now(),read_only=False)
    sheets = fixe_data.active

    sheets[n][0].value = "None"
    sheets[n][1].value = "None"
    sheets[n][2].value = "None"
    sheets[n][3].value = "None"

    fixe_data.save(xlms_path_now())
    fixe_data.close()

# функция для автоматического расчёта опыта
# нужна чтобы понять, сколько опыта можно получить с крафта определённого количества предметов
# доделать нужно! 

# def lvl_setter(BaseXp,Base,Rate):
    